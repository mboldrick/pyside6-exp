import openpyxl
import pandas as pd
import datetime
from mapping_2023 import *


# Prepare a list to collect workout data rows from Excel worksheet
workout_data = []


def open_workbook(path):
    wb = openpyxl.load_workbook(filename=path, data_only=True)
    return wb


def worksheet_exists(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return True
    else:
        return False


def collect_column_names():
    column_names = [
        key
        for key, value in globals().items()
        if type(value) == int and not key.startswith("__")
    ]
    return column_names


def collect_workout_data(wb, sheet_name):
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=5, values_only=True):
        # filtered_row = [cell for index, cell in enumerate(row) if index not in exclude_columns]
        if type(row[DATE]) == datetime.datetime:
            workout_data.append(row)


def main():
    # Open workbook
    filename = "data/Ironman workouts test.xlsx"
    wb = open_workbook(filename)

    # Print summary information about workbook
    print(f"Workbook name: {filename}")
    print(f"Number of worksheets: {len(wb.sheetnames)}")
    # print(f"Worksheets: {wb.sheetnames}")

    # Loop through worksheets (2001 - 2024, exluding 2002, 2004, as they don't exist)
    # for sheet_name in range(2001, 2025):
    #     print(f"Processing worksheet: {sheet_name}", end="")
    #     if worksheet_exists(wb, str(sheet_name)):
    #         print(" - Exists.")
    #     else:
    #         print(" - Does not exist.")

    # Loop through data worksheets
    for sheet_name in range(2023, 2024):
        print(f"Processing worksheet: {sheet_name}", end="")
        if not worksheet_exists(wb, str(sheet_name)):
            print(" - Does not exist.")
        else:
            print(" - Exists.")
            # Get the column names for this worksheet
            colnames = collect_column_names()
            print(f"Number of columns: {len(colnames)}")
            print(f"Columns: {colnames}")
            # Get the worksheet data
            collect_workout_data(wb, str(sheet_name))
            print(f"Number of rows collected: {len(workout_data)}")
            # print(f"Number of entries in first row: {len(workout_data[0])}")
            # Create a dataframe from the workout data
            df = pd.DataFrame(workout_data, columns=colnames)
            # Print the first and last 10 rows of the dataframe
            print(df.head(10))
            print(df.tail(10))
            # Print the dataframe info
            print(df.info())
            # Store the dataframe in a CSV file
            df.to_csv(f"data/{sheet_name}.csv", index=False)
            # Store the dataframe in an Excel file
            df.to_excel(f"data/{sheet_name}.xlsx", index=False)


if __name__ == "__main__":
    main()
