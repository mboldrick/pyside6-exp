import openpyxl
import pandas as pd
import datetime
from mapping_2023 import *


# Define column names
colnames = [
    "week",
    "day_of_week",
    "date",
    "swim_time",
    "swim_meters",
    "swim_hr",
    "bike_time",
    "bike_miles",
    "bike_hr",
    "run_time",
    "run_miles",
    "run_hr",
    "pushups_1",
    "pushups_2",
    "pushups_3",
    "pushups_4",
    "pushups_5",
    "pushups_total",
    "abs_time_1",
    "abs_time_2",
    "abs_time_3",
    "abs_time_total",
    "core_time",
    "meditate_time",
    "strength_time_total",
    "combined_time_total",
    "rest_pulse",
    "bp",
    "sleep_time",
    "weight",
    "body_fat",
    "total_inches",
    "swim_time_total",
    "swim_meters_total",
    "bike_time_total",
    "bike_miles_total",
    "run_time_total",
    "run_miles_total",
    "strength_time_total",
    "combined_time_total",
    "notes",
    "nutrition",
    "meals",
]


def iterating_over_values(path, sheet_name):
    wb = openpyxl.load_workbook(filename=path, data_only=True)

    # List worksheets in workbook
    # for sheet_name in wb.sheetnames:
    #     print(f"{sheet_name}")

    if sheet_name not in wb.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return

    # Process worksheet
    sheet = wb[sheet_name]
    print(f"{sheet_name}")
    print(f"Total number of rows: {sheet.max_row}")
    print(f"Total number of columns: {sheet.max_column}\n")

    # Process individual cells
    # for value in sheet.iter_rows(
    #     min_row=5, max_row=8, min_col=1, max_col=5, values_only=True
    # ):

    #     for cell in value:
    #         print(cell, type(cell))
    #         if type(cell) == datetime.datetime:
    #             print(cell.year)

    weights = []

    for row in sheet.iter_rows(min_row=5, values_only=True):
        # print(row[DATE-1], type(row[DATE-1]))
        if type(row[DATE - 1]) == datetime.datetime:
            weights.append(row)
        else:
            break

    print(f"Number of weights: {len(weights)}")
    print(f"weights[0] variable type: {type(weights[0])}")
    print(weights[0])
    print(weights[0][2], weights[0][25])
