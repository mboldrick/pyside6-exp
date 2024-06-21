from decouple import config
import os
import logging
import openpyxl
from pdf2image import convert_from_path
from PIL import Image, ImageSequence, ImageFilter
from subprocess import PIPE, Popen
import shutil
import uuid
import time
from subprocess import PIPE
import pathlib
import os.path
from os import path
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
from operator import itemgetter
import tifftools
import re
import json
import xml.etree.ElementTree as gfg
from pathlib import Path
import pytesseract
from openpyxl import load_workbook
import copy
import csv
from collections import deque
from collections import OrderedDict
import operator

pytesseract.pytesseract.tesseract_cmd = config("TERRASACT")

unique_id = uuid.uuid4().hex
import pdb

pdb.set_trace()
pid = str(os.getpid())
timestmp = time.strftime("%Y%m%d%H%M%S")
log_formatter = logging.Formatter(
    "%(asctime)s %(levelname)s %(funcName)s line - (%(lineno)d) %(message)s",
    datefmt="%d/%m/%Y %H:%M:%S",
)
# Log File
GlobalProcessFolder = os.path.join(
    config("OUTPUT_TEMP_FOLDER"), "temp_{0}_{1}".format(unique_id, timestmp)
)
logFile = os.path.join(
    config("OUTPUT_LOG_FOLDER"),
    "Preprocessor_log_{0}_{1}.log".format(unique_id, timestmp),
)
detailsfile = os.path.join(
    config("OUTPUT_LOG_FOLDER"),
    "Preprocessor_details_{0}_{1}.xlsx".format(unique_id, timestmp),
)
errorsfile = os.path.join(
    config("OUTPUT_LOG_FOLDER"),
    "Preprocessor_errors_{0}_{1}.csv".format(unique_id, timestmp),
)
tobereviewedfile = os.path.join(
    config("OUTPUT_LOG_FOLDER"),
    "Preprocessor_ToBeReviewed_{0}_{1}.csv".format(unique_id, timestmp),
)
error_preprocessor_trg_file = os.path.join(
    config("ERROR_TRG_FOLDER"), "error_preprocessor.txt"
)
# Setup File handler
file_handler = logging.FileHandler(logFile)
file_handler.setFormatter(log_formatter)
file_handler.setLevel(logging.INFO)
# Setup Stream Handler (i.e. console)
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(log_formatter)
stream_handler.setLevel(logging.INFO)
# Get our logger
logger = logging.getLogger("root")
logger.setLevel(logging.INFO)
# Add both Handlers
logger.addHandler(file_handler)
logger.addHandler(stream_handler)


def sort_form_keys_values(prefixes, values):
    prefixes = sorted(list(set(prefixes)), key=len, reverse=True)
    result_dict = {}
    for prefix in prefixes:
        with_prefix = [x for x in values if x.startswith(prefix)]
        with_prefix = sorted(list(set(with_prefix)), key=len, reverse=True)
        result_dict[prefix] = with_prefix
    return result_dict


def sort_kofax_keys_values(prefixes, data_dict):
    prefixes = sorted(list(set(prefixes)), key=len, reverse=True)
    result_dict = {}
    for prefix in prefixes:
        with_prefix = [x["value"] for x in data_dict if x["key"].startswith(prefix)]
        with_prefix = list(set(with_prefix))
        result_dict[prefix] = with_prefix
    return result_dict


def form_as_dict(loc):
    wrkbk = openpyxl.load_workbook(loc, data_only=True)
    sh = wrkbk.active
    prefixes = []
    values = []
    for i in range(2, sh.max_row + 1):
        prefix = sh.cell(row=i, column=1).value.upper().strip()
        value = sh.cell(row=i, column=2).value.upper().strip()
        prefixes.append(prefix.strip())
        values.append(value)
    return sort_form_keys_values(prefixes, values)


def kofax_as_dict(loc):
    wrkbk = openpyxl.load_workbook(loc, data_only=True)
    sh = wrkbk.active
    prefixes = []
    data_dict = []
    for i in range(2, sh.max_row + 1):
        prefix = sh.cell(row=i, column=1).value.upper().strip()
        try:
            value = sh.cell(row=i, column=2).value.strftime("%Y-%m-%d")
        except:
            continue
        value = sh.cell(row=i, column=2).value
        prefixes.append(prefix)
        data_dict.append({"key": prefix, "value": value})
    return sort_kofax_keys_values(prefixes, data_dict)


def dateValidator(dates, form_text):
    for date in dates:
        date_matched = None
        found = False
        year = date.strftime("%Y")
        month = date.strftime("%m")
        day = date.strftime("%d")
        month_full = date.strftime("%B")
        month_short = date.strftime("%b")

        desired_format = date.strftime("%Y-%m-%d")
        # create some date formarts to search
        format1 = "{0}-{1}-{2}".format(month, day, year[2:])
        format2 = "{0}-{1}-{2}".format(month, day, year)
        format3 = "{0}{1}{2}".format(month, day, year)
        format4 = "{0}{1}".format(month, year)
        format5 = "{0}-{1}".format(month, year)
        format6 = "{0}/{1}".format(month, year)
        format7 = "{0}-{1}".format(month, year[2:])
        format8 = "{0}/{1}/{2}".format(month, day, year)
        format9 = "{0}/{1}/{2}".format(month, day, year[2:])
        format10 = "{0}-{1}".format(month_full, year[2:])
        format11 = "{0}-{1}".format(month_full, year)
        format12 = "{0}-{1}".format(month_short, year[2:])
        format13 = "{0}-{1}".format(month_short, year)
        format14 = "{0}{1}".format(month, year[2:])
        format15 = "{0}{1}".format(month_full, year[2:])
        format16 = "{0}{1}".format(month_full, year)
        format17 = "{0}{1}".format(month_short, year[2:])
        format18 = "{0}{1}".format(month_short, year)
        format19 = "{0}{1}".format(month[1:], year)
        format20 = "{0}{1}".format(month, year[2:])
        format21 = "{0}/{1}".format(month, year[2:])
        format22 = "{0}".format(year)
        format23 = "{0}{1}".format(month[1:], year[2:])
        format24 = "{0}/{1}".format(month[1:], year[2:])

        # you can add other formats below
        potential_date_orders = [
            format1,
            format2,
            format3,
            format4,
            format5,
            format6,
            format7,
            format8,
            format9,
            format10,
            format11,
            format12,
            format13,
            format14,
            format15,
            format16,
            format17,
            format18,
            format19,
            format20,
            format21,
            format22,
            format23,
            format24,
        ]

        for potential_date in potential_date_orders:
            if potential_date in form_text:
                found = True
                #  the desired date format
                date_matched = desired_format
                break
        if found:
            break
    return date_matched, found


def createErrorFile():
    if not os.path.isfile(error_preprocessor_trg_file):
        f = open(error_preprocessor_trg_file, "x")
        f.close()


def deleteErrorFileIfExists():
    if os.path.isfile(error_preprocessor_trg_file):
        os.remove(error_preprocessor_trg_file)


def readsitemapping():
    wrkbk = openpyxl.load_workbook(config("INPUT_SITEMAPPING_EXCEL"), data_only=True)
    sh = wrkbk.active
    data = []
    for i in range(2, sh.max_row + 1):
        officeid = sh.cell(row=i, column=1).value
        agency = sh.cell(row=i, column=2).value
        site_name = sh.cell(row=i, column=3).value
        if site_name is not None:
            data.append(
                {
                    "officeid": officeid,
                    "agency": agency.strip(),
                    "sitename": site_name.strip(),
                }
            )
    return data


def formatText(text):
    text = str(text).lower().strip()
    text = re.sub(r"\W+", "", text.lower())
    return text


def getSiteName(officeid, agency, sitemapping):
    text = ""
    sitefound = False
    for data in sitemapping:
        if formatText(data["officeid"]) == formatText(officeid) and formatText(
            data["agency"]
        ) == formatText(agency):
            text = data["sitename"]
            sitefound = True
            break
    return sitefound, text


def createNewFormPath():
    unique_id = uuid.uuid4().hex
    pid = str(os.getpid())
    timestmp = time.strftime("%Y%m%d%H%M%S")
    return os.path.join(
        config("OUTPUT_LOG_FOLDER"),
        "Preprocessor_details_{0}_{1}.csv".format(unique_id, timestmp),
    )


def write2Csv(data_rows, excelFile):
    with open(excelFile, mode="a", newline="", encoding="utf-8") as form_file:
        form_writer = csv.writer(
            form_file, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL
        )
        for j in range(0, len(data_rows)):
            form_writer.writerow(data_rows[j])


def write2ErrorCsv(data_rows, errorsfilepath):
    if not os.path.isfile(errorsfilepath):
        # deques are optimized for modifications from both endpoints
        data_rows = deque(data_rows)
        data_rows.appendleft(
            ["Input_Path", "FileName", "Agency", "OfficeSiteId", "SiteName"]
        )

    with open(errorsfilepath, mode="a", newline="", encoding="utf-8") as form_file:
        form_writer = csv.writer(
            form_file, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL
        )
        for j in range(0, len(data_rows)):
            form_writer.writerow(data_rows[j])


def office_agency_searchwords_as_dict(office_agency_excel_location):
    wrkbk = openpyxl.load_workbook(office_agency_excel_location, data_only=True)
    sh = wrkbk.active
    data_dict = {}
    for i in range(1, sh.max_row + 1):
        officesiteid = sh.cell(row=i, column=1).value
        agency = sh.cell(row=i, column=2).value.strip()
        data_dict[officesiteid] = agency
    return data_dict


def SiteIdSearch(text, mydict):
    isPresent = False
    agency = None
    officesiteid = None
    for k, v in mydict.items():
        if str(k) in text:
            agency = v
            isPresent = True
            officesiteid = k
            break
    return isPresent, agency, officesiteid


def create_folder(folder):
    Path(folder).mkdir(parents=True, exist_ok=True)


def deleteTemp():
    if os.path.isdir(GlobalProcessFolder):
        shutil.rmtree(GlobalProcessFolder)


def readexcel(input_file):
    wrkbk = openpyxl.load_workbook(input_file, data_only=True)
    sh = wrkbk.active
    d = []
    for i in range(2, sh.max_row + 1):
        dms_doc_type_name = sh.cell(row=i, column=1).value
        dms_doc_type_version = sh.cell(row=i, column=2).value
        dms_doc_type_version = dms_doc_type_version
        d.append(
            {
                "dms_doc_type_name": dms_doc_type_name,
                "dms_doc_type_version": dms_doc_type_version,
            }
        )
    return d


def getfiles(input_folder):
    return [
        str(filepath.absolute())
        for filepath in pathlib.Path(input_folder).glob("**/*")
        if path.isfile(str(filepath.absolute()))
        and (
            str(filepath.absolute()).endswith(".tif")
            or str(filepath.absolute()).endswith(".tiff")
            or str(filepath.absolute()).endswith(".pdf")
        )
    ]


def normalizePath(path):
    return os.path.normpath(path).replace(os.sep, "/")


def extractText(data):
    isPdf = data["isPdf"]
    filepath = data["value"]
    text = None

    if data["index"] == 0:
        if isPdf:
            img = Image.open(filepath)
            CROP_SIZE = 100
            width, height = img.size
            img = img.crop(
                (CROP_SIZE, CROP_SIZE, width - CROP_SIZE, height - CROP_SIZE)
            )
            img = img.resize((600, 800))
            text = pytesseract.image_to_string(img)[:500].strip()
        else:
            img = Image.open(filepath)
            text = pytesseract.image_to_string(img)[:500].strip()
    else:
        img = Image.open(filepath)
        img = img.filter(ImageFilter.DETAIL)
        text = pytesseract.image_to_string(img)[:300].strip()
    img.close()

    text = re.sub(r"[\n\t\r]*", "", text)
    has_separator = False
    if "separator sheet" in text:
        has_separator = True
    is_index = False
    is_second = False
    content = None
    if data["index"] == 0:
        is_index = True
        content = text
    if data["index"] == 1:
        is_second = True

    return {
        "index": data["index"],
        "has_separator": has_separator,
        "extracted_text": text,
        "filepath": filepath,
        "is_index": is_index,
        "content": content,
        "is_second": is_second,
    }


def write_to_xml(fileName, data):
    root = gfg.Element("ImportSession")
    root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    batches = gfg.Element("Batches")
    root.append(batches)
    tree = gfg.ElementTree(root)
    batch = gfg.SubElement(batches, "Batch")
    batch.set("BatchClassName", "MainProcess")
    folders = gfg.SubElement(batch, "Folders")
    folder = gfg.SubElement(folders, "Folder")
    folder.set("FolderClassName", "MainProcess")
    indexfields = gfg.SubElement(folder, "IndexFields")
    documents = gfg.SubElement(folder, "Documents")

    for k, v in data["info"].items():
        indexfield = gfg.SubElement(indexfields, "IndexField")
        indexfield.set("Value", v)
        indexfield.set("Name", k)

    document = gfg.SubElement(documents, "Document")
    pages = gfg.SubElement(document, "Pages")
    for k, v in data["imports"].items():
        page = gfg.SubElement(pages, "Page")
        page.set("ImportFileName", os.path.basename(v))
    with open(fileName, "wb") as files:
        tree.write(files)


def determine_version(index_data_dict):
    keys = index_data_dict.keys()
    has_retention = "retentionperiod" in keys
    has_officeid = "officesiteid" in keys
    has_casenumber = "casenumber" in keys
    version = ""
    if has_retention:
        if len(index_data_dict["retentionperiod"]) > 0:
            version = "1"
    else:
        if has_officeid and not has_casenumber:
            if len(index_data_dict["officesiteid"]) > 0:
                version = "2"
        elif has_officeid and has_casenumber:
            if (
                len(index_data_dict["officesiteid"]) > 0
                and len(index_data_dict["casenumber"]) > 0
            ):
                version = "3"
    return version


def determine_formset(formset):
    mapping = {"kofax": "g1", "missing": "g2", "untrained": "g3", "unindexed": "g4"}
    return mapping[formset]


def determine_agency(agencyStr):
    agencies = ["fsa", "nrcs", "fbc", "nncrs"]

    curr_agency = agencyStr.lower()
    agencyfound = False

    for agency in agencies:
        if agency == curr_agency:
            curr_agency = agency
            agencyfound = True
            break

    return agencyfound, curr_agency


def replacespaces(string):
    if string is not None:
        return "-".join(string.split())
    return None


def processForm(waiting_list):
    try:
        site_name = waiting_list["site_name"]
        agency_name = waiting_list["agency_name"]
        fileguid = waiting_list["fileguid"]
        objectguid = waiting_list["objectguid"]
        original_filename = waiting_list["original_filename"]
        index_data_dict = waiting_list["index_data_dict"]
        matched_date = waiting_list["matched_date"]
        matched_name = waiting_list["matched_name"]
        indexsheetsf = os.path.normpath(waiting_list["indexsheetsf"])
        input_file = os.path.normpath(waiting_list["input_file"])
        original_input_file = os.path.normpath(waiting_list["original_input_file"])
        output_subfolder = os.path.normpath(waiting_list["output_subfolder"])
        version = index_data_dict["indexsheetversion"]
        formset = determine_formset(waiting_list["type"])
        site_id = index_data_dict["officesiteid"]

        agency_subfolder = os.path.normpath(
            os.path.join(indexsheetsf, "{0}".format(agency_name))
        )
        create_folder(agency_subfolder)
        indexSheetGlobal = os.path.join(
            agency_subfolder, "{0}_pp_{1}_IndexSheet.pdf".format(site_id, fileguid)
        )
        if not os.path.exists(indexSheetGlobal):
            commands = ["magick", "convert"]
            commands.append(waiting_list["images"][0])
            commands.append(indexSheetGlobal)
            process = Popen(commands, stdout=PIPE, stderr=PIPE, shell=True)
            process.communicate()

        if waiting_list["has_forms"]:
            data = {
                "info": {
                    "previous_FilePath": "",
                    "previous_JsonFilePath": "",
                    "previous_scansheet": indexSheetGlobal,
                    "Original_OriginalFilePath": original_input_file,
                    "previous_DocumentID": "",
                    "previous_KTAJobID": "",
                    "previous_FormSet": formset,
                    "previous_File_Guid": fileguid,
                    "previous_Object_Guid": objectguid,
                    "Resend": "True",
                },
                "imports": {
                    "ImportIndexFileName": "",
                    "ImportPdfFileName": "",
                },
            }

            matched_name_formatted = replacespaces(matched_name)
            if waiting_list["type"] == "kofax":

                data["info"]["Resend"] = "False"
                toKofaxObjectSubfolder = os.path.normpath(
                    os.path.join(output_subfolder, "toKofax")
                )
                create_folder(toKofaxObjectSubfolder)
                combinedFormPathOutput = os.path.join(
                    toKofaxObjectSubfolder,
                    "{0}_pp_{1}_{2}_{3}_{4}_{5}.pdf".format(
                        site_id,
                        fileguid,
                        objectguid,
                        formset,
                        matched_name_formatted,
                        matched_date,
                    ),
                )
                xml_path = os.path.join(
                    toKofaxObjectSubfolder,
                    "{0}_pp_{1}_{2}_{3}_{4}_{5}.xml".format(
                        site_id,
                        fileguid,
                        objectguid,
                        formset,
                        matched_name_formatted,
                        matched_date,
                    ),
                )
                data["imports"]["ImportPdfFileName"] = combinedFormPathOutput
                logger.info("Combining forms")
                commands = ["magick", "convert"]
                commands.extend(waiting_list["images"][2:])
                commands.append("-adjoin")
                commands.append(combinedFormPathOutput)
                process = Popen(commands, stdout=PIPE, stderr=PIPE, shell=True)
                process.communicate()
                indexSheetGlobalCopy = os.path.join(
                    toKofaxObjectSubfolder,
                    "{0}_pp_{1}_{2}_IndexSheet.pdf".format(
                        site_id, fileguid, objectguid
                    ),
                )
                shutil.copy(indexSheetGlobal, indexSheetGlobalCopy)
                data["imports"]["ImportIndexFileName"] = indexSheetGlobalCopy
                write_to_xml(xml_path, data)
                resolved_formname = combinedFormPathOutput
            else:

                process_sub_folder = os.path.normpath(
                    os.path.join(output_subfolder, "postProcessing")
                )
                snow_agency_folder = os.path.normpath(
                    os.path.join(
                        output_subfolder,
                        "toSnowball/{0}".format(index_data_dict["agency"]),
                    )
                )
                create_folder(snow_agency_folder)
                create_folder(process_sub_folder)

                if waiting_list["type"] == "unindexed":
                    indexSheetGlobalCopy = os.path.join(
                        process_sub_folder,
                        "{0}_pp_{1}_{2}_{3}_Supporting-Material-IndexSheet.pdf".format(
                            site_id, fileguid, objectguid, formset
                        ),
                    )
                    shutil.copy(indexSheetGlobal, indexSheetGlobalCopy)
                    combinedFormPathOutputProcess = os.path.join(
                        process_sub_folder,
                        "{0}_pp_{1}_{2}_{3}_Supporting-Material.pdf".format(
                            site_id, fileguid, objectguid, formset
                        ),
                    )
                    combinedFormPathOutputSnow = os.path.join(
                        snow_agency_folder,
                        "{0}_pp_{1}_{2}_{3}_Supporting-Material.pdf".format(
                            site_id, fileguid, objectguid, formset
                        ),
                    )
                    xml_path = os.path.join(
                        process_sub_folder,
                        "{0}_pp_{1}_{2}_{3}_Supporting-Material.xml".format(
                            site_id, fileguid, objectguid, formset
                        ),
                    )
                    output_json = os.path.join(
                        snow_agency_folder,
                        "{0}_pp_{1}_{2}_{3}_Supporting-Material.json".format(
                            site_id, fileguid, objectguid, formset
                        ),
                    )
                else:
                    indexSheetGlobalCopy = os.path.join(
                        process_sub_folder,
                        "{0}_pp_{1}_{2}_{3}_{4}_IndexSheet.pdf".format(
                            site_id,
                            fileguid,
                            objectguid,
                            formset,
                            matched_name_formatted,
                        ),
                    )
                    shutil.copy(indexSheetGlobal, indexSheetGlobalCopy)
                    combinedFormPathOutputProcess = os.path.join(
                        process_sub_folder,
                        "{0}_pp_{1}_{2}_{3}_{4}.pdf".format(
                            site_id,
                            fileguid,
                            objectguid,
                            formset,
                            matched_name_formatted,
                        ),
                    )
                    combinedFormPathOutputSnow = os.path.join(
                        snow_agency_folder,
                        "{0}_pp_{1}_{2}_{3}_{4}.pdf".format(
                            site_id,
                            fileguid,
                            objectguid,
                            formset,
                            matched_name_formatted,
                        ),
                    )
                    xml_path = os.path.join(
                        process_sub_folder,
                        "{0}_pp_{1}_{2}_{3}_{4}.xml".format(
                            site_id,
                            fileguid,
                            objectguid,
                            formset,
                            matched_name_formatted,
                        ),
                    )
                    output_json = os.path.join(
                        snow_agency_folder,
                        "{0}_pp_{1}_{2}_{3}_{4}.json".format(
                            site_id,
                            fileguid,
                            objectguid,
                            formset,
                            matched_name_formatted,
                        ),
                    )

                data["imports"]["ImportIndexFileName"] = indexSheetGlobalCopy
                data["imports"]["ImportPdfFileName"] = combinedFormPathOutputProcess
                data["info"]["previous_JsonFilePath"] = output_json
                write_to_xml(xml_path, data)
                resolved_formname = combinedFormPathOutputSnow

                template = None

                if waiting_list["type"] == "missing":
                    template = {
                        "ScansheetFields": {
                            "Agency": "",
                            "Binder": "",
                            "Disposition": "",
                            "DispositionDate": "",
                            "Doc": "",
                            "DocTitle": "",
                            "DocType": "",
                            "FileLocation": "",
                            "FolderID": "",
                            "FilePath": "",
                            "FolderTab": "",
                            "LitigationHold": "",
                            "Program": "",
                            "ProgramYear": "",
                            "RetentionPeriod": "",
                            "IndexSheetVersion": "USDAFPAC_IndexSheetv3",
                            "Version": version,
                        },
                        "DocumentFields": {},
                        "DocumentTableFields": {},
                        "DocTypeFields": {
                            "DocTypeName": matched_name,
                            "dms_doc_type_name": matched_name,
                            "dms_doc_type_id": "{0}:".format(matched_name),
                            "dms_doc_type_version": "",
                            "FormSet": formset,
                            "media_type": "forms",
                        },
                        "Filedetails": {
                            "FilePath": "",
                            "CoversheetPath": normalizePath(indexSheetGlobal),
                            "OriginalFilePath": normalizePath(original_input_file),
                            "DocumentID": "",
                            "KTAJobID": "",
                            "Resent": False,
                            "JsonFilePath": normalizePath(output_json),
                            "FileGuid": fileguid,
                            "ObjectGuid": objectguid,
                        },
                    }

                    try:
                        template["ScansheetFields"]["Agency"] = index_data_dict[
                            "agency"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Binder"] = index_data_dict[
                            "binder"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Disposition"] = index_data_dict[
                            "disposition"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["FileLocation"] = index_data_dict[
                            "filelocation"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["FolderTab"] = index_data_dict[
                            "foldertab"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["ProgramYear"] = index_data_dict[
                            "programyear"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Program"] = index_data_dict[
                            "program"
                        ]
                    except:
                        pass

                elif waiting_list["type"] == "untrained":
                    template = {
                        "ScansheetFields": {
                            "ActiveInactive": "",
                            "Agency": "",
                            "AvailableDigitally": "",
                            "Binder": "",
                            "Box": "",
                            "Disposition": "",
                            "DispositionDate": "",
                            "Doc": "",
                            "DocTitle": "",
                            "DocType": "",
                            "FileLocation": "",
                            "FolderID": "",
                            "FolderTab": "",
                            "LitigationHold": "",
                            "OfficeSiteID": "",
                            "Program": "",
                            "ProgramYear": "",
                            "IndexSheetVersion": "USDAFPAC_IndexSheetv2",
                            "Version": version,
                        },
                        "DocumentFields": {},
                        "DocumentTableFields": {},
                        "DocTypeFields": {
                            "DocTypeName": matched_name,
                            "dms_doc_type_name": matched_name,
                            "dms_doc_type_id": "{0}:".format(matched_name),
                            "dms_doc_type_version": matched_date,
                            "FormSet": formset,
                            "media_type": "forms",
                        },
                        "Filedetails": {
                            "FilePath": "",
                            "CoversheetPath": normalizePath(indexSheetGlobal),
                            "OriginalFilePath": normalizePath(original_input_file),
                            "DocumentID": "",
                            "KTAJobID": "",
                            "Resent": False,
                            "JsonFilePath": normalizePath(output_json),
                            "FileGuid": fileguid,
                            "ObjectGuid": objectguid,
                        },
                    }

                    try:
                        template["ScansheetFields"]["OfficeSiteID"] = index_data_dict[
                            "officesiteid"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Doc"] = index_data_dict["doc"]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["ActiveInactive"] = index_data_dict[
                            "activeinactive"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["AvailableDigitally"] = (
                            index_data_dict["availabledigitally"]
                        )
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Box"] = index_data_dict["box"]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Agency"] = index_data_dict[
                            "agency"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Binder"] = index_data_dict[
                            "binder"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Disposition"] = index_data_dict[
                            "disposition"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["FileLocation"] = index_data_dict[
                            "filelocation"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["FolderTab"] = index_data_dict[
                            "foldertab"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["ProgramYear"] = index_data_dict[
                            "programyear"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Program"] = index_data_dict[
                            "program"
                        ]
                    except:
                        pass

                    with open(output_json, "w", encoding="utf-8") as f:
                        json.dump(template, f, ensure_ascii=False, indent=4)

                elif waiting_list["type"] == "unindexed":
                    template = {
                        "ScansheetFields": {
                            "ActiveInactive": "",
                            "Agency": "",
                            "AvailableDigitally": "",
                            "Binder": "",
                            "Box": "",
                            "CaseNumber": "",
                            "Disposition": "",
                            "DispositionDate": "",
                            "Doc": "",
                            "DocTitle": "",
                            "DocType": "",
                            "Drawer": "",
                            "FileCode": "",
                            "FileLocation": "",
                            "FolderTab": "",
                            "LitigationHold": "",
                            "IndexSheetVersion": "USDAFPAC_IndexSheetv3",
                            "OfficeSiteID": "",
                            "Program": "",
                            "ProgramYear": "",
                            "Room": "",
                            "TempPerm": "",
                            "Version": version,
                        },
                        "DocumentFields": {},
                        "DocumentTableFields": {},
                        "DocTypeFields": {
                            "DocTypeName": "Supporting-Material",
                            "dms_doc_type_name": "Supporting-Material",
                            "dms_doc_type_id": "",
                            "dms_doc_type_version": "",
                            "FormSet": formset,
                            "media_type": "forms",
                        },
                        "Filedetails": {
                            "FilePath": "",
                            "CoversheetPath": normalizePath(indexSheetGlobal),
                            "OriginalFilePath": normalizePath(original_input_file),
                            "DocumentID": "",
                            "KTAJobID": "",
                            "Resent": False,
                            "JsonFilePath": normalizePath(output_json),
                            "FileGuid": fileguid,
                            "ObjectGuid": objectguid,
                        },
                    }

                    try:
                        template["ScansheetFields"]["CaseNumber"] = index_data_dict[
                            "casenumber"
                        ]
                    except:
                        pass

                    try:
                        template["ScansheetFields"]["Drawer"] = index_data_dict[
                            "drawer"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Room"] = index_data_dict["room"]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["TempPerm"] = index_data_dict[
                            "tempperm"
                        ]
                    except:
                        pass

                    try:
                        template["ScansheetFields"]["OfficeSiteID"] = index_data_dict[
                            "officesiteid"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Doc"] = index_data_dict["doc"]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["ActiveInactive"] = index_data_dict[
                            "activeinactive"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["AvailableDigitally"] = (
                            index_data_dict["availabledigitally"]
                        )
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Box"] = index_data_dict["box"]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Agency"] = index_data_dict[
                            "agency"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Binder"] = index_data_dict[
                            "binder"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Disposition"] = index_data_dict[
                            "disposition"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["FileLocation"] = index_data_dict[
                            "filelocation"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["FolderTab"] = index_data_dict[
                            "foldertab"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["ProgramYear"] = index_data_dict[
                            "programyear"
                        ]
                    except:
                        pass
                    try:
                        template["ScansheetFields"]["Program"] = index_data_dict[
                            "program"
                        ]
                    except:
                        pass

                commands = ["magick", "convert"]
                commands.extend(waiting_list["images"][2:])
                commands.append("-adjoin")
                commands.append(combinedFormPathOutputProcess)
                process = Popen(commands, stdout=PIPE, stderr=PIPE, shell=True)
                process.communicate()
                combinedFormPathOutputSnowCopy = shutil.copy(
                    combinedFormPathOutputProcess, combinedFormPathOutputSnow
                )

                # updated template
                template["Filedetails"]["FilePath"] = normalizePath(
                    combinedFormPathOutputSnowCopy
                )

                with open(output_json, "w", encoding="utf-8") as f:
                    json.dump(template, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(e)


def getallfiles(input_folder):
    return [
        str(filepath.absolute())
        for filepath in pathlib.Path(input_folder).glob("**/*")
        if path.isfile(str(filepath.absolute()))
    ]


def extractFilename(input_file):
    basename = os.path.basename(input_file)
    return os.path.splitext(basename)[0]


def Image_to_pdf(file: str) -> str:
    filepath = os.path.join(file)
    if filepath.endswith(".tiff"):
        imagepath = filepath.replace(".tiff", ".pdf")
    elif filepath.endswith("tif"):
        imagepath = filepath.replace(".tif", ".pdf")
    if not os.path.exists(filepath):
        raise Exception(f"{filepath} Do not exist.")
    image = Image.open(filepath)
    images = []
    for i, page in enumerate(ImageSequence.Iterator(image)):
        page = page.convert("RGB")
        images.append(page)
    if len(images) == 1:
        images[0].save(imagepath)
    else:
        images[0].save(imagepath, save_all=True, append_images=images[1:])
    image.close()
    if imagepath:
        logger.info(
            "PDF file {0} successfully created".format(extractFilename(imagepath))
        )
        os.remove(filepath)


def searchwords_as_dict(kofaxdb_excel_location):
    wrkbk = openpyxl.load_workbook(kofaxdb_excel_location, data_only=True)
    sh = wrkbk.active
    data_dict = {}
    for i in range(2, sh.max_row + 1):
        curr_prefix = sh.cell(row=i, column=1).value.upper()
        curr_data = sh.cell(row=i, column=2).value.upper()
        if curr_prefix in data_dict.keys():
            l = data_dict[curr_prefix]
            l.append(curr_data)
            data_dict[curr_prefix] = l
        else:
            data_dict[curr_prefix] = [curr_data]

    for k, v in data_dict.items():
        data_dict[k] = sorted(v, reverse=True)

    return data_dict


def text_search_finder(data_list, prefix, raw_text):
    found = False
    text = None
    if prefix in raw_text:
        for data_item in data_list:
            if data_item in raw_text:
                found = True
                text = data_item
                break
    return found, text


def readexcel_searchwords(data_dict, raw_text):
    raw_text = raw_text.upper()
    found = False
    text = None
    for prefix, data_list in data_dict.items():
        found, text = text_search_finder(data_list, prefix, raw_text)
        if found:
            break
    return found, text


def sort_by_value(mydict):
    return {k: v for k, v in sorted(mydict.items(), key=lambda item: item[1])}


def index_to_json_converter(index_input_str):
    data_dict_mapping = {
        "agency": "",
        "binder": "",
        "doctitle": "",
        "doctype": "",
        "doc": "",
        "filelocation": "",
        "folderid": "",
        "foldertab": "",
        "litigationnahold": "",
        "programyear": "",
        "program": "",
        "retentionperiod": "",
        "activeinactive": "",
        "availabledigitally": "",
        "box": "",
        "dispositiondate": "",
        "disposition": "",
        "officesiteid": "",
        "casenumber": "",
        "drawer": "",
        "filecode": "",
        "indexsheetversion": "",
        "room": "",
        "tempperm": "",
        "holdtype": "",
        "retention1period": "",
        "pyear": "",
        "ddate": "",
        "dtitle": "",
        "dtype": "",
    }

    index_input_str = index_input_str.replace("programyear", "pyear")
    index_input_str = index_input_str.replace("dispositiondate", "ddate")
    index_input_str = index_input_str.replace("doctype", "dtype")
    index_input_str = index_input_str.replace("doctitle", "dtitle")

    index_data_as_dict = {}
    for x in data_dict_mapping.keys():
        index = index_input_str.find(x)

        if index != -1:
            index_data_as_dict[x] = index

    sorted_values = sort_by_value(index_data_as_dict)
    i = 0
    p = 0
    l = None
    d = {}
    for k, v in sorted_values.items():
        start = p
        end = v
        if i > 0:
            d[l] = index_input_str[start:end]
        p = v + len(k)
        l = k
        i += 1

    d[l] = index_input_str[p:]

    try:
        d["retentionperiod"] = d["retention1period"]
    except:
        pass
    try:
        d["programyear"] = d["pyear"]
    except:
        pass
    try:
        d["dispositiondate"] = d["ddate"]
    except:
        pass
    try:
        d["doctype"] = d["dtype"]
    except:
        pass
    try:
        d["doctitle"] = d["dtitle"]
    except:
        pass

    del data_dict_mapping["retention1period"]
    del data_dict_mapping["pyear"]
    del data_dict_mapping["ddate"]
    del data_dict_mapping["dtype"]
    del data_dict_mapping["dtitle"]

    for k, v in d.items():
        data_dict_mapping[k] = v
    return data_dict_mapping


def getUUID():
    return uuid.uuid4().hex


def process_files(
    input_file,
    file_count,
    working_folder,
    Kofax_DB_dict,
    searchwords,
    agency_officeid_hints,
    sitemapping,
    fileguid,
):
    Agency_G = None
    OfficeSiteId_G = None
    SiteName_G = None
    isPdf = input_file.endswith(".pdf")
    formdetails = {"isSucess": False, "content": []}

    original_filename = extractFilename(input_file)
    original_input_file = input_file

    try:
        # 1. TODO. save to buffer instead of disk
        logger.info("Splitting images...")
        start_time = time.time()
        if isPdf:
            logger.info("converting pdf into images...")
            try:
                images = convert_from_path(input_file, poppler_path=config("POPPLER"))
                for i in range(len(images)):
                    images[i].save(
                        os.path.join(working_folder, "page_{0}.tiff".format(i)), "JPEG"
                    )
                    images[i].close()
            except Exception as e:
                logger.info(e)
                raise Exception(e)
        else:
            logger.info("converting tiff into images...")
            try:
                currentFileName = input_file
                info = tifftools.read_tiff(currentFileName)
                for i, ifd in enumerate(info["ifds"]):
                    tifftools.write_tiff(
                        ifd, os.path.join(working_folder, "page_{0}.tiff".format(i))
                    )
            except Exception as e:
                logger.info(e)
                raise Exception(e)

        seconds_taken = time.time() - start_time
        logger.info("--- Splitting took %s seconds to complete ---" % seconds_taken)

        # 2. get the location of separator sheet with multithreading(scanning each page)
        # 3. save first 300 characters of each page to array for future, thus prevent second scan
        logger.info("Fetching all splitted images in : {0}".format(working_folder))
        logger.info("Detecting separator sheet and saving first 300 chars")
        logger.info("Extracting Text from the Images...")
        splittedFiles = getallfiles(working_folder)
        splittedFiles = sorted(
            splittedFiles,
            key=lambda i: int(os.path.splitext(os.path.basename(i))[0].split("_")[-1]),
        )

        operation_list = []

        for index in range(0, len(splittedFiles)):
            operation_list.append(
                {
                    "index": index,
                    "value": splittedFiles[index],
                    "file_count": (file_count + 1),
                    "isPdf": isPdf,
                }
            )

        separator_hints = []
        extracted_text = []
        new_files = []
        index_input_str = None

        extraction_error = False
        exception = None
        with ThreadPoolExecutor(max_workers=350) as executor:
            future_to_filepath = {
                executor.submit(extractText, data): data for data in operation_list
            }
            result = []
            for future in concurrent.futures.as_completed(future_to_filepath):
                try:
                    result.append(future.result())
                except Exception as e:
                    logger.info(e)
                    exception = e
                    extraction_error = True
                    break
            executor.shutdown()
            new_files.extend(result)

        # if any extraction_error, return early and do nothing
        if extraction_error:
            logger.info(exception)
            return True, formdetails

        sorted_result = sorted(new_files, key=itemgetter("index"))
        invalid_file = False
        for x in sorted_result:
            separator_hints.append(x["has_separator"])
            extracted_text.append(x["extracted_text"])
            if x["is_index"]:
                index_input_str = x["content"]
            if x["is_second"] and not x["has_separator"]:
                invalid_file = True
        if invalid_file:
            raise Exception("Invalid File")

        logger.info(index_input_str)
        index_input_str = re.sub(r"\W+", "", index_input_str.lower())
        index_data_dict = index_to_json_converter(index_input_str.lower())

        Agency_G = index_data_dict["agency"]
        OfficeSiteId_G = index_data_dict["officesiteid"]

        has_agency = len(index_data_dict["agency"]) != 0
        has_officeid = len(index_data_dict["officesiteid"]) != 0

        if not has_agency and not has_officeid:
            isPresent, agency, officesiteid = SiteIdSearch(
                index_input_str, agency_officeid_hints
            )
            if isPresent:
                index_data_dict["agency"] = agency
                index_data_dict["officesiteid"] = officesiteid
            else:
                raise Exception("Either oOfficeSiteID or Agency is missing.")

        Tobereviewed = False

        if not has_agency or not has_officeid:
            Tobereviewed = True
            logger.info(
                "Unable to extract Agency or OfficeSiteID from tesseract output"
            )

        agencyfound, agency_name = determine_agency(index_data_dict["agency"])

        sitefound = False
        site_name = None
        if has_officeid:
            sitefound, site_name = getSiteName(
                index_data_dict["officesiteid"], agency_name, sitemapping
            )

        if not agencyfound or not sitefound:
            Tobereviewed = True
            logger.info("Agency or OfficeSiteID mismatch")

        SiteName_G = site_name
        if not Tobereviewed:
            # 4. create combined forms prototype
            acc = []
            files_indexes = []
            if len(separator_hints) > 2:
                for i in range(2, len(separator_hints)):
                    if separator_hints[i]:
                        if len(acc) > 0:
                            entry = [0, 1]
                            entry.extend(acc)
                            files_indexes.append(entry)
                        acc = []
                    else:
                        acc.append(i)

            if len(acc) > 0:
                entry = [0, 1]
                entry.extend(acc)
                files_indexes.append(entry)

            if len(files_indexes) == 0 and len(separator_hints) <= 2:
                if len(separator_hints) == 1:
                    files_indexes = [[0]]

            # 5. inputs: (excel data, 300 save in step 3, combined form prototypes, splitted images on disk, i.e will be buffer in future)
            # 6. output: combined form, sorted form i.e unindexed, kofax, missing date
            # 7. prepare for multithreaded image combinations and sorting
            kofaxf = os.path.join(os.path.normpath(config("OUTPUT_FOLDER")), "kofax")
            missingf = os.path.join(
                os.path.normpath(config("OUTPUT_FOLDER")), "MissingEdition"
            )
            unindexedf = os.path.join(
                os.path.normpath(config("OUTPUT_FOLDER")), "Unstructured"
            )
            indexsheetsf = os.path.join(
                os.path.normpath(config("OUTPUT_FOLDER")), "IndexSheets"
            )
            form = os.path.join(os.path.normpath(config("OUTPUT_FOLDER")), "Untrained")

            output_folders = [kofaxf, missingf, unindexedf, form, indexsheetsf]

            for f in output_folders:
                create_folder(f)

            waiting_list = []
            data_list = []
            form_count = 0

            for current_indexes in files_indexes:
                start_time = time.time()
                name_only = False
                name_time = False
                isform = False
                matched_word = None
                matched_date = None
                matched_name = None
                has_forms = False
                newDate = None

                if len(current_indexes) > 2:
                    target = current_indexes[2]
                    form_text = extracted_text[target]
                    form_text = "".join(
                        char for char in form_text if char.isalnum() or char in {"-"}
                    )
                    logger.info(form_text)
                    logger.info("\n" * 2)
                    has_forms = True
                    Kofax_DB_dict = {
                        k.replace(" ", ""): v for k, v in Kofax_DB_dict.items()
                    }

                    with_name_matches = []
                    for name, dates in Kofax_DB_dict.items():
                        if name in form_text.upper():
                            with_name_matches.append({"name": name, "dates": dates})
                            matching_date, found = dateValidator(dates, form_text)
                            name_only = True
                            if found:
                                name_time = True
                                matched_name = name
                                matched_date = matching_date
                                break

                    if name_only and not name_time:
                        matched_name = with_name_matches[0]["name"]

                    if name_only == False:
                        found, text = readexcel_searchwords(searchwords, form_text)
                        isform = found
                        matched_word = text
                        matched_name = text

                objectguid = "{0}".format(getUUID())
                form_name = "{0}_{1}.tif".format(original_filename, objectguid)
                form_name_noex = "{0}_{1}".format(
                    fileguid, objectguid
                )  # form name without extension for logging

                default_data = {
                    "output_subfolder": unindexedf,
                    "images": [splittedFiles[x] for x in current_indexes],
                    "form_name": form_name,
                    "original_filename": original_filename,
                    "fileguid": fileguid,
                    "objectguid": objectguid,
                    "type": None,
                    "input_file": input_file,
                    "index_data_dict": index_data_dict,
                    "matched_date": matched_date,
                    "matched_name": matched_name,
                    "indexsheetsf": indexsheetsf,
                    "has_forms": has_forms,
                    "original_input_file": original_input_file,
                    "agency_name": agency_name,
                    "site_name": site_name,
                }

                data_dict = {
                    "filename": original_filename,
                    "Kofax": 0,
                    "MissingEdition": 0,
                    "Unindexed": 0,
                    "Untrained": 0,
                    "Pages": 0,
                    "Form_name": form_name_noex,
                    "Edition_Date": None,
                    "Form_code": None,
                    "Keyword_search": None,
                    "Missing_Edition_name": None,
                    "agency": index_data_dict["agency"],
                    "officesiteid": index_data_dict["officesiteid"],
                    "input_path": original_input_file,
                    "site_name": site_name,
                }

                if len(current_indexes) > 2:
                    data_dict["Pages"] = (
                        len(current_indexes) - 2
                    )  # total length of the form list - separator and Index page
                else:
                    data_dict["Pages"] = 0

                # update output folder
                if name_time:
                    default_data["output_subfolder"] = kofaxf
                    default_data["type"] = "kofax"
                    waiting_list.append(default_data)
                    data_dict["Kofax"] = 1
                    data_dict["Form_code"] = matched_name
                    data_dict["Edition_Date"] = matched_date
                elif name_only:
                    default_data["type"] = "missing"
                    default_data["output_subfolder"] = missingf
                    waiting_list.append(default_data)
                    data_dict["MissingEdition"] = 1
                    data_dict["Missing_Edition_name"] = matched_name
                elif isform:
                    default_data["type"] = "untrained"
                    default_data["output_subfolder"] = form
                    waiting_list.append(default_data)
                    data_dict["Untrained"] = 1
                    data_dict["Keyword_search"] = matched_word
                else:
                    default_data["type"] = "unindexed"
                    default_data["output_subfolder"] = unindexedf
                    waiting_list.append(default_data)
                    data_dict["Unindexed"] = 1

                data_list.append(data_dict)
                form_count += 1

            formdetails["isSucess"] = True
            formdetails["content"] = data_list
            try:
                with ThreadPoolExecutor(350) as executor:
                    executor.map(processForm, waiting_list)
                    executor.shutdown()
            except Exception as e:
                logger.info(e)
                raise Exception(e)

            processed_folder = os.path.normpath(
                os.path.join(config("OUTPUT_FOLDER"), "Processed")
            )
            create_folder(processed_folder)
            final_folder = processed_folder

        else:
            tobereviewed_input_folder = os.path.normpath(
                os.path.join(config("OUTPUT_FOLDER"), "ToBeReviewed")
            )
            create_folder(tobereviewed_input_folder)
            input_file_parent = os.path.normpath(os.path.dirname(input_file)).replace(
                os.sep, "/"
            )
            inp_folder = os.path.normpath(os.path.join(config("INPUT_FOLDER"))).replace(
                os.sep, "/"
            )
            logger.info("sending to be tobereviewed")
            logger.info(input_file_parent)
            logger.info(inp_folder)
            start = input_file_parent.find(inp_folder) + len(inp_folder) + 1
            relative_path = input_file_parent[start:]
            tobereviewed_input_folder_revised = os.path.normpath(
                os.path.join(tobereviewed_input_folder, relative_path)
            ).replace(os.sep, "/")
            create_folder(tobereviewed_input_folder_revised)
            final_folder = tobereviewed_input_folder_revised
            write2ErrorCsv(
                [
                    [
                        original_input_file,
                        os.path.basename(original_input_file),
                        Agency_G,
                        OfficeSiteId_G,
                        SiteName_G,
                    ]
                ],
                tobereviewedfile,
            )

        processed_path = shutil.move(input_file, final_folder)
        # was it an image, change to PDF
        if not processed_path.endswith(".pdf"):
            Image_to_pdf(processed_path)

    except Exception as e:
        logger.info("sending to error folder")
        logger.info(e)
        error_folder = os.path.join(os.path.normpath(config("OUTPUT_FOLDER")), "Error")
        create_folder(error_folder)
        shutil.move(input_file, error_folder)
        write2ErrorCsv(
            [
                [
                    original_input_file,
                    os.path.basename(original_input_file),
                    Agency_G,
                    OfficeSiteId_G,
                    SiteName_G,
                ]
            ],
            errorsfile,
        )
    return False, formdetails


def task():
    # delete start.start if exists
    start_start = os.path.join(config("INPUT_FOLDER"), "start.start")
    if os.path.isfile(start_start):
        os.remove(start_start)

    # delete error file if exists
    if os.path.isfile(error_preprocessor_trg_file):
        os.remove(error_preprocessor_trg_file)

    deleteErrorFileIfExists()
    input_files = getfiles(config("INPUT_FOLDER"))
    formhints_dict = form_as_dict(config("INPUT_FORMHINTS_EXCEL"))
    Kofax_DB_dict = kofax_as_dict(config("INPUT_KOFAXDB_EXCEL"))
    agency_officeid_hints = office_agency_searchwords_as_dict(
        config("INPUT_AGENCY_OFFICEID_EXCEL")
    )
    sitemapping = readsitemapping()

    header_data = {
        "Input_Path": "input_path",
        "FileName": "filename",
        "FormName": "Form_name",
        "Agency": "agency",
        "OfficeSiteId": "officesiteid",
        "SiteName": "site_name",
        "Kofax": "Kofax",
        "MissingEdition": "MissingEdition",
        "Untrained": "Untrained",
        "Unstructured": "Unindexed",
        "G1-Form-ID": "Form_code",
        "G1-EditionDate": "Edition_Date",
        "G2-Form-ID": "Missing_Edition_name",
        "G3-Form-ID": "Keyword_search",
        "Pages": "Pages",
    }

    limit = 2000
    headers = list(header_data.keys())
    index = 0
    inputFilesCount = len(input_files)
    excelFile = None
    if inputFilesCount > 0:
        excelFile = createNewFormPath()
        write2Csv([headers], excelFile)

    workingFileTempFolder = None
    for i in range(0, inputFilesCount):
        try:
            logger.info("processing input file {0}: {1}".format(i + 1, input_files[i]))
            file = extractFilename(input_files[i])
            workingFileTempFolder = os.path.join(GlobalProcessFolder, file)
            create_folder(workingFileTempFolder)
            fileguid = "{0}".format(getUUID())
            canbreak, formdetails = process_files(
                input_files[i],
                i,
                workingFileTempFolder,
                Kofax_DB_dict,
                formhints_dict,
                agency_officeid_hints,
                sitemapping,
                fileguid,
            )
            deleteTemp()

            logger.info(
                "completed processing file {0}: {1}".format(i + 1, input_files[i])
            )
            logger.info("\n" * 3)
            if canbreak:
                logger.info("file extraction error.Terminating the program...")
                logger.info("Deleting Temp folder")
                deleteTemp()
                originaldir = os.path.dirname(input_files[i])
                logger.info("Moving Original file back to Input folder")
                shutil.move(input_file, originaldir)
                logger.info("Create a trigger file to restart the process")
                createErrorFile()
                break
            else:
                if formdetails["isSucess"]:
                    for form in formdetails["content"]:
                        data = []
                        for k, v in header_data.items():
                            data.append(form[v])
                        if index == (limit - 1):
                            write2Csv([data], excelFile)
                            oldfile = excelFile
                            # move old to backup flder
                            backup_folder = os.path.join(
                                os.path.normpath(config("OUTPUT_LOG_FOLDER")),
                                "BackupFolder",
                            )
                            create_folder(backup_folder)
                            shutil.move(oldfile, backup_folder)
                            # create a new one
                            excelFile = createNewFormPath()
                            write2Csv([headers], excelFile)
                            index = 0
                        else:
                            write2Csv([data], excelFile)
            index += 1
        except Exception as e:
            logger.info("Error while processing the file.Terminating the program...")
            logger.info(e)
            logger.info("Deleting Temp folder")
            deleteTemp()
            logger.info("Create a trigger file to restart the process")
            createErrorFile()


if __name__ == "__main__":
    logger.info("program started running")
    start_time = time.time()
    task()
    logger.info("program completed running")
    seconds_taken = time.time() - start_time
    logger.info("--- program excecuted in %s seconds ---" % seconds_taken)
